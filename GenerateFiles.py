import glob
import os
from genericpath import isdir, isfile
import os
import sys
import shutil
import glob
from datetime import datetime
from ctypes import windll
import string
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def ExcelTesting():
    wb = Workbook()
    dest_filename = 'empty_book.xlsx'
    ws1 = wb.active
    ws1.title = "range names"
    for row in range(1, 40):
        ws1.append(range(600))
    ws2 = wb.create_sheet(title="Pi")
    ws2['F5'] = 3.14
    ws3 = wb.create_sheet(title="Data")
    for row in range(10, 20):
        for col in range(27, 54):
            _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
    print(ws3['AA10'].value)
    wb.save(filename = dest_filename)

def get_drives():
    drives = []
    bitmask = windll.kernel32.GetLogicalDrives()
    for letter in string.ascii_uppercase:
        if bitmask & 1:
            drives.append(letter)
        bitmask >>= 1
    return drives

def GenerateListOfFilesInDir(dir_name):
# Get a list of files (file paths) in the given directory 
    list_of_files = filter( os.path.isfile,
                        glob.glob(dir_name + '/**/*', recursive=True) )
# get list of ffiles with size
    files_with_size = [ (file_path, os.stat(file_path).st_size) 
                    for file_path in list_of_files ]
    return files_with_size
    
def GetListOfDirsInCurrentDrive(drive_Path):
    list_subfolders_with_paths = [f.path for f in os.scandir(drive_Path) if f.is_dir()]
    #print(list_subfolders_with_paths)
    return list_subfolders_with_paths
# Iterate over list of tuples i.e. file_paths with size
# and print them one by one
if __name__ == "__main__":
    argCount = len(sys.argv)
    print('Argument List: ', str(sys.argv) , argCount)
    #ExcelTesting()
    driveLetters = get_drives()
    book = Workbook()
    dest_filename = 'ListOfFiles.xlsx'
    for driveLetter in driveLetters:
        drivePath = driveLetter + ":\\"
        WorkSheetName = driveLetter + " Drive"
        if (not WorkSheetName in book.get_sheet_names()):
            book.create_sheet(WorkSheetName,0)
        curWS = book[WorkSheetName]
        curWS.column_dimensions['A'].width = 120
        curWS.column_dimensions['B'].width = 70
        pathCol,nameCol,sizeCol,rowIndex = 1, 2 , 3, 1
        for dir_path in GetListOfDirsInCurrentDrive(drivePath):
            listOfSystemDirs = ["Program Files","Program Files (x86)","Microsoft","Windows", "ProgramData"]
            if not any(system_dir in dir_path for system_dir in listOfSystemDirs):
                _ = curWS.cell(column=pathCol, row=rowIndex, value=dir_path)
                rowIndex+=1
                print(dir_path)
                for fileData in GenerateListOfFilesInDir(dir_path):
                    _ = curWS.cell(column=pathCol, row=rowIndex, value=os.path.dirname(fileData[0]))
                    _ = curWS.cell(column=nameCol, row=rowIndex, value=os.path.basename(fileData[0]))
                    _ = curWS.cell(column=sizeCol, row=rowIndex, value=fileData[1])
                    rowIndex+=1
    book.save(filename = dest_filename)
                 
    # if (argCount != 2):
        # print("Error Please provide valid Argument")
        # exit()
    # drivePath = sys.argv[1]
    